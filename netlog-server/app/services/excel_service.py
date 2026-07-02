# excel_service.py
# 검수 기록을 Excel(.xlsx) 파일로 생성하는 서비스 모듈.
# collection.py의 /inspection/export 엔드포인트에서 호출된다.

import io
from datetime import date, datetime, timezone, timedelta

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text

# DB에 저장된 시각은 UTC 기준 TIMESTAMPTZ이므로, 엑셀 출력 시 KST(+9)로 변환한다.
KST = timezone(timedelta(hours=9))

# ── 스타일 상수 ──────────────────────────────────────────────────────────────
# 헤더 행: 파란 배경(#1E88E5) + 흰 굵은 글씨
HEADER_FILL  = PatternFill(fgColor="1E88E5", fill_type="solid")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 메타 정보 행(집하장명, 조회 기간): 굵은 글씨
META_FONT    = Font(bold=True, size=10)

# 홀수 데이터 행 배경: 연한 파란색(#F0F7FF)으로 줄무늬 효과
ROW_EVEN_FILL = PatternFill(fgColor="F0F7FF", fill_type="solid")

# 모든 데이터 셀에 적용할 얇은 테두리
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

# 수거 상태별 글자 색상 (수거완료: 진파랑, 부분수거: 주황, 그 외: 회색)
STATUS_COLORS = {
    "수거완료": "1565C0",
    "부분수거": "E65100",
    "수거대기": "757575",
    "대기중":   "757575",
}

# 엑셀 헤더 행 레이블 (A~H열 순서)
HEADERS = ["검수일시", "집하장명", "선박명", "마대 수량 (자루)", "수거 완료 여부", "잔여 마대 수", "이미지 URL", "검수 ID"]


def _fmt_dt(dt) -> str:
    """UTC(또는 naive) datetime을 KST 기준 'YYYY-MM-DD HH:MM' 문자열로 변환한다."""
    if dt is None:
        return ""
    try:
        # tzinfo가 없는 naive datetime은 UTC로 간주하고 KST로 변환
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(KST).strftime("%Y-%m-%d %H:%M")
    except Exception:
        # 변환 실패 시 원본 문자열 그대로 반환 (데이터 손실 방지)
        return str(dt)


def _derive_collection_status(row: dict) -> tuple[str, int]:
    """
    site_bag_queue 테이블의 수거 정보를 바탕으로 수거 상태 문자열과 잔여 수량을 계산한다.

    반환값: (상태 문자열, 잔여 마대 수)
      - site_bag_queue 행 없음(LEFT JOIN NULL) → ("대기중", 원래 수량)
      - is_fully_collected = True              → ("수거완료", 0)
      - remaining < original                  → ("부분수거", remaining)
      - remaining == original                 → ("수거대기", remaining)
    """
    # site_bag_queue에 해당 record_id 행이 없는 경우 (LEFT JOIN 결과가 NULL)
    if row["is_fully_collected"] is None:
        return "대기중", row["bag_count"]

    if row["is_fully_collected"]:
        return "수거완료", 0

    remaining = row["remaining_bag_count"]
    original  = row["original_bag_count"]

    if remaining < original:
        return "부분수거", remaining

    return "수거대기", remaining


def _get_inspection_records_for_export(db, site_id: str, start_date: date, end_date: date, vessel_id_list: list = []) -> list[dict]:
    start_dt = datetime(start_date.year, start_date.month, start_date.day, 0, 0, 0, tzinfo=KST)
    end_dt   = datetime(end_date.year,   end_date.month,   end_date.day,   0, 0, 0, tzinfo=KST) + timedelta(days=1)

    where_clauses = [
        "ir.site_id = :site_id",
        "ir.inspected_at >= :start_dt",
        "ir.inspected_at <  :end_dt"
    ]
    params = {"site_id": site_id, "start_dt": start_dt, "end_dt": end_dt}

    # 선박 필터 추가
    if vessel_id_list:
        placeholders = ",".join([f":vid_{i}" for i in range(len(vessel_id_list))])
        where_clauses.append(f"ir.vessel_id::text IN ({placeholders})")
        for i, vid in enumerate(vessel_id_list):
            params[f"vid_{i}"] = vid

    where_sql = " AND ".join(where_clauses)

    rows = db.execute(
        text(f"""
            SELECT
                ir.record_id,
                ir.inspected_at,
                ir.bag_count,
                ir.bag_image_url,
                s.name  AS site_name,
                v.name  AS vessel_name,
                q.is_fully_collected,
                q.remaining_bag_count,
                q.original_bag_count
            FROM inspection_record ir
            JOIN site   s ON s.site_id   = ir.site_id
            JOIN vessel v ON v.vessel_id = ir.vessel_id
            LEFT JOIN site_bag_queue q ON q.record_id = ir.record_id
            WHERE {where_sql}
            ORDER BY ir.inspected_at ASC
        """),
        params
    ).fetchall()

    return [row._mapping for row in rows]


def generate_inspection_excel(db, site_id: str, start_date: date, end_date: date, vessel_id_list: list = []) -> bytes:
    """
    검수 기록 Excel 파일을 생성하여 bytes로 반환한다.
    반환된 bytes는 라우터에서 StreamingResponse에 직접 전달된다.

    시트 구조:
      행 1 : 집하장명 (메타 정보)
      행 2 : 조회 기간 (메타 정보)
      행 3 : 빈 행 (시각적 구분)
      행 4 : 헤더 (파란 배경)
      행 5~ : 데이터 (홀수 행에 연한 파란 배경 적용)
    """
    records = _get_inspection_records_for_export(db, site_id, start_date, end_date, vessel_id_list)

    # 데이터가 없을 경우 site_name을 빈 문자열로 처리 (메타 행은 그대로 출력)
    site_name = records[0]["site_name"] if records else ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "검수 기록"

    # ── 메타 정보 행 (1~3행) ─────────────────────────────────────────────────
    ws.append(["집하장명", site_name])
    ws["A1"].font = META_FONT
    ws.append(["조회 기간", f"{start_date} ~ {end_date}"])
    ws["A2"].font = META_FONT
    ws.append([])  # 헤더와의 시각적 구분을 위한 빈 행

    # ── 헤더 행 (4행) ────────────────────────────────────────────────────────
    ws.append(HEADERS)
    for col_idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border    = THIN_BORDER

    # ── 데이터 행 (5행~) ─────────────────────────────────────────────────────
    for row_num, row in enumerate(records, start=5):
        # site_bag_queue 정보로 수거 상태 및 잔여 수량 계산
        status, remaining = _derive_collection_status(dict(row))

        ws.cell(row=row_num, column=1).value = _fmt_dt(row["inspected_at"])  # A: 검수일시 (KST)
        ws.cell(row=row_num, column=2).value = row["site_name"]              # B: 집하장명
        ws.cell(row=row_num, column=3).value = row["vessel_name"]            # C: 선박명
        ws.cell(row=row_num, column=4).value = row["bag_count"]              # D: 마대 수량

        # E: 수거 완료 여부 — 상태별로 글자 색상을 다르게 표시
        status_cell       = ws.cell(row=row_num, column=5)
        status_cell.value = status
        status_cell.font  = Font(color=STATUS_COLORS.get(status, "000000"))

        ws.cell(row=row_num, column=6).value = remaining                     # F: 잔여 마대 수

        # G: 이미지 URL — 클릭 가능한 하이퍼링크로 표시 ("이미지 보기" 텍스트)
        image_cell = ws.cell(row=row_num, column=7)
        if row["bag_image_url"]:
            image_cell.value     = "이미지 보기"
            image_cell.hyperlink = row["bag_image_url"]
            image_cell.font      = Font(color="1565C0", underline="single")

        ws.cell(row=row_num, column=8).value = str(row["record_id"])         # H: 검수 ID (UUID)

        # 홀수 데이터 행(5, 7, 9...)에 연한 파란 배경을 적용하여 줄무늬 효과
        if row_num % 2 == 1:
            for col_idx in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                if col_idx != 5 and col_idx != 7:  # 상태/하이퍼링크 셀은 font가 별도 지정되므로 fill만 적용
                    cell.fill = ROW_EVEN_FILL
                elif col_idx == 5:
                    cell.fill = ROW_EVEN_FILL

        # 모든 데이터 셀에 얇은 테두리 적용
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=row_num, column=col_idx).border = THIN_BORDER

    # ── 컬럼 너비 자동 조정 ───────────────────────────────────────────────────
    # 각 열에서 가장 긴 셀 값의 글자 수 기준으로 너비를 설정한다. (최대 50)
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 50)

    # 워크북을 메모리 버퍼에 저장하고 bytes로 반환 (서버 디스크에 파일을 쓰지 않음)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
